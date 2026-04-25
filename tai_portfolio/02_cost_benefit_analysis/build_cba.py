"""
EMMA System Cost-Benefit Analysis Model
Temple Allen Industries – Business Development Portfolio
Author: Yadnesh Deshpande
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ────────────────────────────────────────────────────────────────────

def hdr(bold=True, size=11, color="FFFFFF"):
    return Font(bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def ctr():
    return Alignment(horizontal="center", vertical="center")

def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

NAVY   = "1F3864"
BLUE   = "2E75B6"
LGRAY  = "D9E1F2"
WHITE  = "FFFFFF"
GREEN  = "E2EFDA"
YELLOW = "FFF2CC"

# ── Sheet 1: Assumptions ──────────────────────────────────────────────────────

ws1 = wb.active
ws1.title = "Assumptions"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 22

# Title
ws1.merge_cells("A1:C1")
ws1["A1"] = "EMMA™ System – Cost-Benefit Analysis"
ws1["A1"].font = Font(bold=True, size=14, color=WHITE)
ws1["A1"].fill = fill(NAVY)
ws1["A1"].alignment = ctr()

ws1.merge_cells("A2:C2")
ws1["A2"] = "Customer: Aerospace MRO Facility  |  Prepared by: Yadnesh Deshpande  |  Date: April 2026"
ws1["A2"].font = Font(size=9, color="666666")
ws1["A2"].fill = fill("F2F2F2")
ws1["A2"].alignment = ctr()

ws1.row_dimensions[1].height = 28
ws1.row_dimensions[2].height = 16

# Section headers
sections = [
    (4,  "INVESTMENT",           [
        ("EMMA System Purchase Price ($)",          350000),
        ("Installation & Setup ($)",                 12000),
        ("Training ($)",                              8000),
        ("Annual Maintenance ($)",                   18000),
        ("Consumables per Year ($)",                  6500),
    ]),
    (11, "CURRENT STATE (Manual)", [
        ("Number of Manual Sanding Workers",              6),
        ("Average Worker Hourly Rate ($/hr)",            28),
        ("Hours Worked per Day",                          8),
        ("Working Days per Year",                       250),
        ("Rework Rate (% of jobs)",                    0.12),
        ("Average Rework Cost per Job ($)",             420),
        ("Jobs per Year",                               800),
        ("Worker Injury Rate (incidents/yr)",             2),
        ("Average Injury Cost ($)",                   15000),
    ]),
    (21, "EMMA PERFORMANCE",      [
        ("EMMA Throughput vs Manual (%)",              2.30),
        ("EMMA Rework Rate (%)",                       0.02),
        ("EMMA Operator Headcount (from 6 to)",           1),
        ("EMMA Uptime (%)",                            0.95),
    ]),
]

row = 3
for start_row, title, params in sections:
    ws1.row_dimensions[start_row].height = 20
    ws1.merge_cells(f"A{start_row}:C{start_row}")
    ws1[f"A{start_row}"] = title
    ws1[f"A{start_row}"].font = hdr(color=WHITE)
    ws1[f"A{start_row}"].fill = fill(BLUE)
    ws1[f"A{start_row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for i, (label, val) in enumerate(params):
        r = start_row + 1 + i
        ws1.row_dimensions[r].height = 17
        ws1[f"A{r}"] = label
        ws1[f"A{r}"].font = Font(size=10)
        ws1[f"A{r}"].fill = fill(LGRAY if i % 2 == 0 else WHITE)

        ws1[f"B{r}"] = val
        ws1[f"B{r}"].font = Font(size=10, color="0000FF", bold=True)
        ws1[f"B{r}"].fill = fill(YELLOW)
        ws1[f"B{r}"].alignment = Alignment(horizontal="right")
        if isinstance(val, float) and val < 1:
            ws1[f"B{r}"].number_format = "0.0%"
        elif isinstance(val, int) and val > 100:
            ws1[f"B{r}"].number_format = "$#,##0"

        ws1[f"C{r}"] = "← Input (blue = editable)"
        ws1[f"C{r}"].font = Font(size=9, color="888888", italic=True)

        for col in ["A", "B", "C"]:
            ws1[f"{col}{r}"].border = thin_border()

# ── Sheet 2: CBA Model ────────────────────────────────────────────────────────

ws2 = wb.create_sheet("CBA Model")
ws2.sheet_view.showGridLines = False
for col, w in zip(["A","B","C","D","E","F","G"], [32,14,14,14,14,14,14]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells("A1:G1")
ws2["A1"] = "EMMA™ Cost-Benefit Analysis – 5-Year Projection"
ws2["A1"].font = Font(bold=True, size=13, color=WHITE)
ws2["A1"].fill = fill(NAVY)
ws2["A1"].alignment = ctr()
ws2.row_dimensions[1].height = 26

# Year headers
ws2["A2"] = "Item"
ws2["A2"].font = hdr(color=WHITE)
ws2["A2"].fill = fill(BLUE)
ws2["A2"].alignment = Alignment(horizontal="left", indent=1)

for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
    col = get_column_letter(i + 2)
    ws2[f"{col}2"] = yr
    ws2[f"{col}2"].font = hdr(color=WHITE)
    ws2[f"{col}2"].fill = fill(BLUE)
    ws2[f"{col}2"].alignment = ctr()

ws2.row_dimensions[2].height = 18

# Sections
rows_data = [
    ("COSTS", NAVY, None),
    ("System Purchase",          "$#,##0", ["=Assumptions!B5",  0,  0,  0,  0]),
    ("Installation & Setup",     "$#,##0", ["=Assumptions!B6",  0,  0,  0,  0]),
    ("Training",                 "$#,##0", ["=Assumptions!B7",  0,  0,  0,  0]),
    ("Annual Maintenance",       "$#,##0", ["=Assumptions!B8", "=Assumptions!B8", "=Assumptions!B8", "=Assumptions!B8", "=Assumptions!B8"]),
    ("Consumables",              "$#,##0", ["=Assumptions!B9", "=Assumptions!B9", "=Assumptions!B9", "=Assumptions!B9", "=Assumptions!B9"]),
    ("Total Annual Cost",        "$#,##0", ["=SUM(B4:B8)", "=SUM(C4:C8)", "=SUM(D4:D8)", "=SUM(E4:E8)", "=SUM(F4:F8)"]),
    ("SAVINGS", NAVY, None),
    ("Labor Savings",            "$#,##0", ["=(Assumptions!B12-Assumptions!B23)*Assumptions!B13*Assumptions!B14*Assumptions!B15"]*5),
    ("Rework Cost Reduction",    "$#,##0", ["=(Assumptions!B16-Assumptions!B24)*Assumptions!B17*Assumptions!B18"]*5),
    ("Injury Cost Avoided",      "$#,##0", ["=Assumptions!B19*Assumptions!B20"]*5),
    ("Total Annual Savings",     "$#,##0", ["=SUM(B10:B12)", "=SUM(C10:C12)", "=SUM(D10:D12)", "=SUM(E10:E12)", "=SUM(F10:F12)"]),
    ("NET BENEFIT", NAVY, None),
    ("Net Benefit (Savings–Cost)","$#,##0",["=B13-B9","=C13-C9","=D13-D9","=E13-E9","=F13-F9"]),
    ("Cumulative Net Benefit",   "$#,##0", ["=B15","=B15+C15","=B15+C15+D15","=B15+C15+D15+E15","=B15+C15+D15+E15+F15"]),
    ("ROI (%)",                  "0.0%",   ["=B13/B9","=C13/C9","=D13/D9","=E13/E9","=F13/F9"]),
]

r = 3
for item in rows_data:
    label, fmt, vals = item
    ws2.row_dimensions[r].height = 18

    if vals is None:
        ws2.merge_cells(f"A{r}:G{r}")
        ws2[f"A{r}"] = label
        ws2[f"A{r}"].font = hdr(color=WHITE, size=10)
        ws2[f"A{r}"].fill = fill(BLUE)
        ws2[f"A{r}"].alignment = Alignment(horizontal="left", indent=1)
    else:
        is_total = label.startswith("Total") or label.startswith("Net") or label.startswith("Cumul") or label.startswith("ROI")
        bg = "D9E1F2" if is_total else ("F2F2F2" if r % 2 == 0 else WHITE)

        ws2[f"A{r}"] = label
        ws2[f"A{r}"].font = Font(bold=is_total, size=10)
        ws2[f"A{r}"].fill = fill(bg)
        ws2[f"A{r}"].alignment = Alignment(indent=1)

        for i, v in enumerate(vals):
            col = get_column_letter(i + 2)
            cell = ws2[f"{col}{r}"]
            cell.value = v if v != 0 else "-"
            cell.font = Font(bold=is_total, size=10, color="000000")
            cell.fill = fill(GREEN if is_total else bg)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            cell.border = thin_border()

        ws2[f"A{r}"].border = thin_border()

    r += 1

# ── Sheet 3: Summary Dashboard ────────────────────────────────────────────────

ws3 = wb.create_sheet("Executive Summary")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 36
ws3.column_dimensions["B"].width = 22

ws3.merge_cells("A1:B1")
ws3["A1"] = "EMMA™ Investment – Executive Summary"
ws3["A1"].font = Font(bold=True, size=13, color=WHITE)
ws3["A1"].fill = fill(NAVY)
ws3["A1"].alignment = ctr()
ws3.row_dimensions[1].height = 26

kpis = [
    ("Total 5-Year Investment",    "='CBA Model'!B9+'CBA Model'!C9+'CBA Model'!D9+'CBA Model'!E9+'CBA Model'!F9", "$#,##0"),
    ("Total 5-Year Savings",       "='CBA Model'!B13+'CBA Model'!C13+'CBA Model'!D13+'CBA Model'!E13+'CBA Model'!F13", "$#,##0"),
    ("5-Year Net Benefit",         "='CBA Model'!F16", "$#,##0"),
    ("Payback Period (approx yrs)","=('CBA Model'!B9)/('CBA Model'!B13)", "0.0"),
    ("5-Year Average ROI",         "=AVERAGE('CBA Model'!B17,'CBA Model'!C17,'CBA Model'!D17,'CBA Model'!E17,'CBA Model'!F17)", "0.0%"),
    ("Labor Headcount Reduction",  "=Assumptions!B12-Assumptions!B23", "0 workers"),
]

for i, (label, formula, fmt) in enumerate(kpis):
    r = i + 3
    ws3.row_dimensions[r].height = 22
    ws3[f"A{r}"] = label
    ws3[f"A{r}"].font = Font(size=11)
    ws3[f"A{r}"].fill = fill(LGRAY if i % 2 == 0 else WHITE)
    ws3[f"A{r}"].border = thin_border()

    ws3[f"B{r}"] = formula
    ws3[f"B{r}"].font = Font(bold=True, size=11, color="1F3864")
    ws3[f"B{r}"].fill = fill(GREEN)
    ws3[f"B{r}"].number_format = fmt
    ws3[f"B{r}"].alignment = Alignment(horizontal="center")
    ws3[f"B{r}"].border = thin_border()

wb.save("/home/claude/tai_portfolio/02_cost_benefit_analysis/EMMA_CBA_Model.xlsx")
print("✓ EMMA_CBA_Model.xlsx created")
